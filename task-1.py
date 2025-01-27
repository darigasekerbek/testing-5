import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pytest
import openpyxl
from webdriver_manager.chrome import ChromeDriverManager


def get_test_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    test_data = []
    
    for row in range(2, sheet.max_row + 1): 
        username = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value
        test_data.append((username, password))
    
    return test_data

def login(username, password):
    options = Options()

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    try:
        print("Открытие сайта...")
        driver.get("https://jut.su")
        driver.maximize_window()
        time.sleep(3)

        login_btn = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//span[text()='Войти']")))
        login_btn.click()

        username_field = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='login_input1']")))
        username_field.send_keys(username)
        time.sleep(3)

        password_field = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//input[@id='login_input2']")))
        password_field.send_keys(password)

        login_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='login_submit']")))
        login_button.click()
        time.sleep(3)

        try:
            exit_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//a[text()='[Выйти]']")))
            print("Вход успешно выполнен!")
        except:
            print("Не удалось войти. Проверьте учетные данные.")

        exit_btn.click()
        time.sleep(2)
        print("Выход успешно выполнен!")

    finally:
        driver.quit()
        print("Браузер закрыт.")

@pytest.mark.parametrize("username, password", get_test_data('testdata.xlsx'))
def test_login(username, password):
    print(f"Тест с параметрами: {username}, {password}")
    login(username, password)
